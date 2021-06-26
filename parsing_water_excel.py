"""
Данный скрипт формирует два списка ежедневных показаний входящего и исходящего
чётчиков горячей воды за прошедший период:
list_in
list_out
"""

from win32com.client import Dispatch
import openpyxl
from variables import \
    days_in_last_month, \
    path_journal, \
    month_before_last_date_YYYYMM as YYYYMM, \
    month_before_last_date_YYYY as YYYY

# открываем книгу excel '_Показания всех счётчиков.xlsm'
wb = openpyxl.load_workbook(filename=r'D:\0=0=0=0=0\ГерПан#\. Счётчики\Показания\_Показания всех счётчиков.xlsm',\
                            read_only=False, keep_vba=True)
# задаём активный лист файла '_Показания всех счётчиков.xlsm'
ws = wb['Main']
# находим максимальную строку в столбце 'C'
max_row = max((C.row for C in ws['C'] if C.value is not None))
# показание счётчика входящей воды на начало расчётного периода
last_in = ws[f'C{max_row-1}'].value
# показание счётчика выходящей воды на начало расчётного периода
last_out = ws[f'D{max_row-1}'].value
# показание счётчика входящей воды на конец расчётного периода
now_in = ws[f'C{max_row}'].value
# показание счётчика выходящей воды на конец расчётного периода
now_out = ws[f'D{max_row}'].value
# расход воды за период
water_rate = (now_in - last_in) - (now_out - last_out)
# закрываем книгу excel '_Показания всех счётчиков.xlsm'
wb.close()

# количество воды проходящей ежедневно через входящий счётчик
everyday_in = (now_in - last_in)/(days_in_last_month - 1)
# количество воды проходящей ежедневно через выходящий счётчик
everyday_out = (now_out - last_out)/(days_in_last_month - 1)

# формируем список ежедневных показаний входящего счётчика
list_in = [last_in]
# чтобы не изменять переменную last_in (необходима для дальнейшей проверки),
# создаём другую переменную
count_in = last_in
for i in range(1, days_in_last_month -1):
    count_in += everyday_in
    list_in.append(round(count_in))
list_in.append(now_in)

# формируем список ежедневных показаний выходящего счётчика
list_out = [last_out]
# чтобы не изменять переменную last_out (необходима для дальнейшей проверки),
# создаём другую переменную
count_out = last_out
for i in range(1, days_in_last_month -1):
    count_out += everyday_out
    list_out.append(round(count_out))
list_out.append(now_out)
#---------------------------------------
### находим последние показания входящего и выходящего счётчика горячей
### воды позапрошлого месяца
# запускаем приложение excel
xl = Dispatch('Excel.Application')
# открываем файл журнала 'YYYYMM' позапрошлого месяца
wb = xl.Workbooks.Open(rf'{path_journal}\{YYYY}\{YYYYMM}')
# определяем активный лист открытого файла
ws = wb.Worksheets('Протокол')
# находим номер последней строки (int)
last_row = ws.UsedRange.Rows.Count
# показание счётчика входящей воды на конец расчётного периода позапрошлого месяца
last_in_old = int(ws.Cells((last_row - 2), 5).Value)
# показание счётчика выходящей воды на конец расчётного периода позапрошлого месяца
last_out_old = int(ws.Cells((last_row - 2), 6).Value)
# закрываем файл 'YYYYMM'
wb.Close()
# закрываем приложение excel
xl.Quit()
#---------------------------------------
# если не совпадают показания счётчиков на начала расчётного периода прошлого месяца
# и конец расчётного периода позапрошлого месяца
if last_in_old != last_in or last_out_old != last_out:
    # формируем списки с ежедневным error
    list_in = ['error'] * days_in_last_month
    list_out = list_in