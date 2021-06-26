"""
Данный скрипт формирует три списка показаний счётчика КМ-5 учёта тепловой энергии:
list_Q
list_M1
list_T
"""

import os
import openpyxl
from variables import days_in_last_month, \
    last_month_date_YYYYMM, \
    diff_dates, \
    desktop, \
    path_journal

# имя последней папки в папке "Журнал"
target_folder_name = os.listdir(path_journal)[-1]
# имя последнего файла в последней папке внутри папки "Журнал"
target_file_name = os.listdir(path_journal + '\\' + target_folder_name)[-1]
# путь к последнему существующему файлу отчёта в папке "Журнал"
target_file_path = path_journal + '\\' + target_folder_name + '\\' + target_file_name


# если существует файл {last_month_date_YYYYMM}.txt на рабочем столе
if os.path.exists(f'{desktop}/{last_month_date_YYYYMM}.txt'):

    # парсим данные по расходу тепловой энергии за прошлый месяц из файла {last_month_date_YYYYMM}.txt
    with open(f'{desktop}/{last_month_date_YYYYMM}.txt', 'r') as file:

        # создаём список из строк текстового файла 'KM5.txt'
        lines = file.readlines()

        # дата начала расчетного периода
        date_start = lines[10][28:38]

        # дата окончания расчетного периода
        date_finish = lines[10][44:54]

        # количество дней мониторинга в файле last_month_date_YYYYMM.txt
        days = diff_dates(date_start, date_finish) + 1

        # парсим количество потребленной тепловой энергии за расчётный период
        total_Q = float(lines[19 + days][14:23].replace(' ', ''))

        # парсим циркуляцию теплоносителя за расчётный период
        total_M1 = float(lines[19 + days][24:32].replace(' ', ''))

        # парсим показание расхода тепла на начало расчётного периода
        start_Q = lines[27 + days][34:44].replace(' ', '')

        # парсим показание расхода тепла на конец расчётного периода
        stop_Q = lines[26 + days][34:44].replace(' ', '')

        # парсим циркуляцию теплоносителя на начало расчётного периода
        start_M1 = lines[27 + days][45:54].replace(' ', '')

        # парсим циркуляцию теплоносителя на конец расчётного периода
        stop_M1 = lines[26 + days][45:54].replace(' ', '')

        # создаём список едеждевного расхода тепла
        daily_usage_Q = []
        for i in range(18, 18 + days):
            daily_usage_Q.append(float(lines[i][14:23].replace(' ', '')))

        # создаём список едеждевной циркуляции теплоносителя
        daily_usage_M1 = []
        for i in range(18, 18 + days):
            daily_usage_M1.append(float(lines[i][24:32].replace(' ', '')))

        # создаем список ежедневных показаний счётчика расхода тепла
        # на нулевое место этого списка ставим значение показания счётчика
        # расхода тепла на конец прошлого периода
        list_Q = [float(start_Q)]
        change_daily_counter = float(start_Q)
        for i in range(1, len(daily_usage_Q) - 1):
            if daily_usage_Q[i] == 0.0:
                change_daily_counter += daily_usage_Q[i-1]
            else:
                change_daily_counter += daily_usage_Q[i]
            list_Q.append(round(change_daily_counter, 3))
        list_Q.append(float(stop_Q))

        # создаем список ежедневных показаний счётчика циркуляции теплоносителя
        # на нулевое место этого списка ставим значение показания счётчика
        # циркуляции теплоносителя на конец прошлого периода
        list_M1 = [float(start_M1)]
        change_daily_counter = float(start_M1)
        for i in range(1, len(daily_usage_M1) - 1):
            if daily_usage_M1[i] == 0.0:
                change_daily_counter += daily_usage_M1[i-1]
            else:
                change_daily_counter += daily_usage_M1[i]
            list_M1.append(round(change_daily_counter, 3))
        list_M1.append(float(stop_M1))

        # парсим время работы приборов учёта и создаём соответствующий список
        list_T = []
        for i in range(18, 18 + days):
            list_T.append(float(lines[i][64:70].replace(' ', '')))

        # если актуальная отчёту дата в формате YYYYMM не совпадает с датой в том же
        # формате из файла YYYYMM.txt или количество отчётных дней из файла YYYYMM.txt
        # не совпадает с количеством дней в прошлом месяце (т.е. этот текстовый файл
        # YYYYMM.txt был сформирован с ошибкой)
        if date_finish[-4:] + date_finish[3:5] != str(last_month_date_YYYYMM) or \
           date_start[-4:]  + date_start[3:5] != str(last_month_date_YYYYMM) or \
           days != days_in_last_month:
            list_Q  = ['error'] * days_in_last_month
            list_M1 = ['error'] * days_in_last_month
            list_T  = ['error'] * days_in_last_month
            start_Q = 'error'
            stop_Q  = 'error'
            total_Q = 'error'
else:
    # открываем последний существующий файл отчёта
    book = openpyxl.load_workbook(filename=target_file_path, read_only=True)
    # начинаем обрабатывать единственную активную страницу последнего файла отчёта
    sheet = book.active
    # создаём списки из последних показаний приборов учёта тепловой энергии
    # с одинаковыми ежедневными показаниями приборов учёта тепловой энергии,
    # т.к. они не работали в данном отчётном периоде
    list_Q = [sheet[sheet.max_row - 2][1].value] * days_in_last_month
    list_M1 = [sheet[sheet.max_row - 2][2].value] * days_in_last_month
    list_T = [0.0] * days_in_last_month
    # задаём значения переменных для заполнения файла справки по расходу теплоэнергии
    start_Q = sheet[sheet.max_row - 2][1].value
    stop_Q = start_Q
    total_Q = '0,000'

# коррекция списка list_Q
list_QQ = list_Q.copy()
for i in range(len(list_Q)-2, 0, -1):
    if list_Q[i] == list_Q[i-1]:
        list_QQ[i] = list_QQ[i + 1]
list_Q = list_QQ

# коррекция списка list_M1
list_MM = list_M1.copy()
for i in range(len(list_M1)-2, 0, -1):
    if list_M1[i] == list_M1[i-1]:
        list_MM[i] = list_MM[i + 1]
list_M1 = list_MM

